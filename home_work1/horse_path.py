import time
from collections import deque

import openpyxl

from helpers import Helper

# for testing only
#from timeit import Timer

horse = '\u265E'  # white horse = '\u2658'
alg_time_field = 'AP2'  # hardcoded

source = 'kv008horse.xlsx'
destination = 'horse008.xlsx'

wb = openpyxl.load_workbook(filename=source)
sheet = wb.active
# top start B4  top end DF4 bottom start B123 bottom end DF123
BOARD_RANGE = 'B4:DF123'
# tuple of style_id where cell is empty
EMPTY_CELLS = (1, 3, 4, 6, 7, 8, 9, 11, 18)
top_border = 3  # sheet.BREAK_ROW
left_border = 1  # sheet.BREAK_COLUMN
rows = len(sheet.rows) - top_border
cols = len(sheet.columns) - left_border
start = (0, 0)  # start point
finish = (119, 108) # end point
helper = Helper(rows, cols, start)
queue = deque()
# parents array with points
parents = dict()
results = []
coordinates = []

matrix = [[0] * cols for i in range(rows)]

for i, row in enumerate(sheet.iter_rows(BOARD_RANGE)):
    for j, cell in enumerate(row):
        if cell.style_id not in EMPTY_CELLS:
            matrix[i][j] = -1

def aka_bfs(start, goal):
    queue.append(start)
    found = False
    while(len(queue) > 0 and not found):
        step = queue.popleft()
        if step == goal:
            found = True
        next_steps = helper.horse_moves(step)
        for nst in next_steps:
            if not parents.get(nst, 0) and matrix[nst[0]][nst[1]] != -1:
                queue.append(nst)
                parents[nst] = step

t0 = time.time()
aka_bfs(start, finish)
t1 = time.time() - t0

# part 3
helper.get_short_path(finish, parents, results)
board_list = list(map(lambda x: (x[0] + 4, x[1] + 2), results))

# _ =sheet.cell(column=col, row=row, value=horse)

for elem in board_list:
    coordinates.append(str(sheet.cell(row=elem[0],
                                      column=elem[1]).coordinate))

for coo in coordinates:
    sheet[coo].value = horse

sheet[alg_time_field].value = t1
wb.save(filename=destination)

# open in exel
# from xlwings import Workbook, Sheet, Range
# wb = Workbook(fullname='kv008horse.xlsx')
# for coo in coordinates:
#    Range(coo).value = horse
