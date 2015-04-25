import time
from collections import deque
import openpyxl


class Horse(object):

    def __init__(self, board):
        self.board = board  # instance of Board class
        self.matrix = self.board.matrix
        self.start = self.board.get_start()
        self.finish = self.board.get_finish()

    def run(self, start=None, finish=None, matrix=None):
        matrix = deck or self.matrix
        start = start or self.start
        finisth = finish or self.finish
        result = []
        moves = self._bfs(start, finish, matrix)
        self.get_short_path(self.goal, moves, result)
        return result

    def _bfs(self, start, goal, matrix):
        queue = deque()
        parents = dict()
        queue.append(start)
        found = False
        while(queue and not found):
            step = queue.popleft()
            if step == goal:
                found = True
            next_steps = self._horse_moves(step)
            for nst in next_steps:
                if not parents.get(nst, 0) and matrix[nst[0]][nst[1]] != -1:
                    queue.append(nst)
                    parents[nst] = step
        return parents

    def _horse_moves(self, point):
        orig = self.board.limit
        possible_places = [(1, -2), (-1, -2), (-2, -1), (-2, 1),
                           (-1, 2), (1, 2), (2, 1), (2, -1)]

        gen_moves = lambda x: (x[0] + point[0], x[1] + point[1])
        is_legal = lambda x: x[0] >= 0 and x[0] < orig[0]
            and x[1] >= 0 and x[1] < orig[1], temp

        return list(filter(is_legal, map(gen_moves, possible_places)))

    def get_short_path(self, finish, parents, result):
        if (parents[finish] != self.start):
            self.get_short_path(parents[finish], parents, result)
        result.append(finish)


class Board(object):
    EMPTY_CELLS = (1, 3, 4, 6, 7, 8, 9, 11, 18)

    def __init__(self, sheet, iterobj, start, finish):
        """ sheet = instance of openpyxl-workbook-active
            board_range = "A1:DF123"
        """
        self.start = start
        self.finish = finish
        self.sheet = sheet
        self.iterobj = iterobj
        left_border = sheet.BREAK_ROW
        top_border = sheet.BREAK_COLUMN + 1
        self.limit = (
            len(sheet.rows) - top_border, len(sheet.columns) - left_border)
        self.matrix = [[0] * self.limit[1] for i in range(self.limit[0])]
        self._fill_board(iterobj, start, finish)

    def _fill_board(self, iterobj, start, finish):
        for i, row in enumerate(iterobj):
            for j, cell in enumerate(row):
                if cell.style_id not in self.EMPTY_CELLS:
                    self.matrix[i][j] = -1
                elif cell.value == start:
                    self.start = (i, j)
                elif cell.value == finish:
                    self.finish = (i, j)

    def new_matrix(self, start_val, finish_val):
        self._fill_board(start_val, finish_val)
        return self.matrix

    def get_matrix():
        return self.matrix

    def get_start(self):
        return self.start

    def get_finish(self):
        return self.finish


def main():
    horse = '\u265E'  # white horse = '\u2658'
    alg_time_field = 'AP2'  # hardcoded
    source = 'kv008horse.xlsx'
    destination = 'horse008.xlsx'
    start = 's'
    finish = 'v'
    BOARD_RANGE = "B4:DF123"

    wb = openpyxl.load_workbook(filename=source)
    sheet = wb.active
    iter_object = sheet.iter_rows(BOARD_RANGE)

    main_board = Board(sheet, tier_object, start, finish)

    horse = Horse(main_board)
    result_steps = horse.run()

    # hardcode
    board_list = list(map(lambda x: (x[0] + 4, x[1] + 2), result_steps))

    # _ =sheet.cell(column=col, row=row, value=horse)
    coordinates = []
    for elem in board_list:
        coordinates.append(str(sheet.cell(row=elem[0],
                                          column=elem[1]).coordinate))

    for coo in coordinates:
        sheet[coo].value = horse

    # sheet[alg_time_field].value = t1
    wb.save(filename=destination)

    # open in exel
    # from xlwings import Workbook, Sheet, Range
    # wb = Workbook(fullname='kv008horse.xlsx')
    # for coo in coordinates:
    #    Range(coo).value = horse

if __name__ == "__main__":
    main()

# https://libxlsxwriter.github.io/hello_8c-example.html
# 